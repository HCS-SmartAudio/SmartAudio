"""
ExcelHandle: Class defined to read and write parameters into an Excel sheet.
"""
import openpyxl
import re


class ExcelHandler(object):
    colheading = ["Steps to be followed", "Expected output/observation", "Results"]
    steptitle = "Steps to be followed"
    stepexpected = "Expected output/observation"
    results = "Results"

    def open_excel_file(self, file_location, sheetname):
        """
        Function defined to get sheet name and work book
        :param file_location:
        :param sheetname:
        :return: sheet name, work book
        """
        workbook = openpyxl.load_workbook(file_location)
        sname = workbook[sheetname]
        return sname, workbook

    def get_col_name_for_steps(self, sname, colheading):
        """
        Function defined to get the column name for deriving steps from the same.
        :param sname:
        :param colheading:
        :return: column name
        """
        cols = sname.max_column
        print(cols)
        for i in range(cols):
            i = i + 1
            if re.search(colheading, str(sname.cell(row=1, column=i).value)):
                return i
                break

    def update_results(self, wb, sname, row, results):
        """
        Function defined to update the results in the excel sheet.
        :param wb:
        :param sname:
        :param row:
        :param results:

        """
        cols = sname.max_column

        for i in range(cols):
            i = i + 1
            if re.search(r"Results", str(sname.cell(row=1, column=i).value)):
                sname.cell(row,i).value = results

    def read_row_data(self,sname):
        """"
        Function defined to read the data from a row
        :param sname
        """
        rows=sname.max_row
        for i in range(rows):
            i=i+2
            if str(sname.cell(row=i, column=5).value) != "None":
                rowdata=sname.cell(i,5).value
                print(rowdata)

    def read_cell_data(self, sname, rownum):
        """"
         Function defined to read the data from a cell
         :param sname rownum
         :return: Cell data
        """
        return str(sname.cell(rownum, column=5).value)

    def get_row_count(self,sname):
        """
        Function defined to get the maximum number of rows
        :param sname:
        :return: maximum number of rows
        """
        return sname.max_row


    def get_column_count(self,sname):
        """
        Function defined to get the maximum number of columns
        :param sname:
        :return: maximum number of columns
        """
        return sname.max_column



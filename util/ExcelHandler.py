
import openpyxl
import re

'''
ExcelHandle: Class defined to read and write parameters into an Excel sheet.
'''

class ExcelHandle:
    colheading = ["Steps to be followed", "Expected output/observation", "Results"]
    steptitle = "Steps to be followed"
    stepexpected = "Expected output/observation"
    results = "Results"

    '''
    open_excel_file : Function defined to open an excel file.
    Parameters In: File location and sheetname
    Parameters Out: sheet name, workbook name
    '''

    def open_excel_file(self, file_location, sheetname):
        workbook = openpyxl.load_workbook(file_location)
        sname = workbook[sheetname]
        return sname, workbook

    '''
    getcolnameforsteps: Function defined to get the column name for deriving steps from the same.
    Parameters In: sheetname and column heading
    Parameters Out: column name
    '''

    def getcolnameforsteps(self, sname, colheading):
        cols = sname.max_column
        print(cols)
        for i in range(cols):
            i = i + 1
            if re.search(colheading, str(sname.cell(row=1, column=i).value)):
                return i
                break

    '''
    update_results: Function defined to update the results in the excel sheet.
    Parameters In: workbook, sheet_name, row no, result
    '''

    def update_results(self, wb, sname, row, results):
        cols = sname.max_column

        for i in range(cols):
            i = i + 1
            if re.search(r"Results", str(sname.cell(row=1, column=i).value)):
                sname.cell(row,i).value = results

    '''
    read_rowdata: Function defined to read the data from a row
    Parameters In: sheetname
    '''

    def read_rowdata(self,sname):
        rows=sname.max_row
        for i in range(rows):
            i=i+2
            if str(sname.cell(row=i, column=5).value) != "None":
                rowdata=sname.cell(i,5).value
                print(rowdata)

    '''
    read_celldata: Function defined to read the data from a cell
    Parameters In: sheetname, row number
    Parameters Out: Cell data
    '''

    def read_celldata(self, sname, rownum):
        return str(sname.cell(rownum, column=5).value)

    '''
    get_row_count: Function defined to get the maximum number of rows
    Parameters In: sheetname
    Parameters Out: maximum number of rows
    '''

    def get_row_count(self,sname):
        return sname.max_row

    '''
    get_column_count: Function defined to get the maximum number of columns
    Parameters In: sheetname
    Parameters Out: maximum number of columns
    '''

    def get_column_count(self,sname):
        return sname.max_column

